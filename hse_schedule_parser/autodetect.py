"""Smart detection of courses, groups, and modules from Excel files.

Opens an .xlsx workbook and scans for:
- Available course sheets (1 курс, 2 курс, etc.)
- Group codes within each course sheet
- Module number and period from the title row
- Subgroup information
- Week parity from the "Календарь" sheet
"""

from __future__ import annotations

import logging
import re
from datetime import date
from pathlib import Path

import openpyxl

from hse_parser.reader import (
    COURSE_PREFIX_MAP,
    COURSE_TO_SHEET,
    _group_exists_in_sheet,
    extract_title_info,
    find_group_columns,
    resolve_sheet_name,
)
from hse_parser.utils import normalize_group_code

logger = logging.getLogger(__name__)

# Month names in nominative case (as they appear in calendar header rows)
CALENDAR_MONTHS: dict[str, int] = {
    "Сентябрь": 9,
    "Октябрь": 10,
    "Ноябрь": 11,
    "Декабрь": 12,
    "Январь": 1,
    "Февраль": 2,
    "Март": 3,
    "Апрель": 4,
    "Май": 5,
    "Июнь": 6,
}

# Module block start rows in the "Календарь" sheet (1-based)
MODULE_BLOCK_ROWS = [7, 18, 29, 40]
# Day-of-week rows within each module block (0-based offset from block start)
DAY_ROW_OFFSETS = [3, 4, 5, 6, 7, 8, 9]  # rows 10-16, 21-27, 32-38, 43-49
# First data column is D (column 4)
FIRST_DATA_COL = 4


def _open_workbook_readonly(file_path: str | Path) -> openpyxl.Workbook | None:
    """Open an .xlsx workbook in read-only mode.

    Returns the workbook, or None if the file doesn't exist or can't be opened.
    """
    file_path = Path(file_path)
    if not file_path.exists():
        return None
    try:
        return openpyxl.load_workbook(str(file_path), read_only=True)
    except Exception:
        return None


def detect_courses(file_path: str | Path) -> list[int]:
    """Detect which course sheets are available in the workbook.

    Returns a sorted list of course numbers (e.g. [1, 2, 3, 4]).
    """
    wb = _open_workbook_readonly(file_path)
    if wb is None:
        return []

    try:
        available_sheets = set(wb.sheetnames)
    finally:
        wb.close()

    courses: list[int] = []
    for course_num, sheet_name in COURSE_TO_SHEET.items():
        if sheet_name in available_sheets:
            courses.append(course_num)
        # Also check "1 курс." variant for 1st year
        if course_num == 1 and "1 курс." in available_sheets:
            if 1 not in courses:
                courses.append(1)

    return sorted(courses)


def detect_groups(file_path: str | Path, course: int) -> list[str]:
    """Detect all group codes available for a given course.

    Returns a sorted list of group code strings.
    """
    wb = _open_workbook_readonly(file_path)
    if wb is None:
        return []

    try:
        sheet_name = resolve_sheet_name(wb, course)
        if sheet_name is None or sheet_name not in wb.sheetnames:
            return []

        ws = wb[sheet_name]
        groups: list[str] = []

        # Scan row 10 for group codes
        for row in ws.iter_rows(min_row=10, max_row=10, values_only=False):
            for cell in row:
                if cell.value is not None:
                    val = str(cell.value).strip()
                    # Match group code pattern: 2 digits + letters
                    if re.match(r"^\d{2}[А-ЯЁа-яё0-9]+$", val):
                        groups.append(val)

        return sorted(set(groups))
    finally:
        wb.close()


def detect_module_info(
    file_path: str | Path, course: int
) -> dict | None:
    """Detect module number and period for a given course.

    Returns dict with 'module', 'period_start', 'period_end' or None.
    """
    wb = _open_workbook_readonly(file_path)
    if wb is None:
        return None

    try:
        sheet_name = resolve_sheet_name(wb, course)
        if sheet_name is None or sheet_name not in wb.sheetnames:
            return None

        ws = wb[sheet_name]
        title_info = extract_title_info(ws)
        if title_info is None:
            return None

        module, period_start, period_end = title_info
        return {
            "module": module,
            "period_start": period_start,
            "period_end": period_end,
        }
    finally:
        wb.close()


def detect_subgroups(
    file_path: str | Path, group_code: str
) -> list[int]:
    """Detect if a group has subgroup split.

    Scans the schedule data for 'гр.1' and 'гр.2' markers.
    Returns list of detected subgroup numbers (e.g. [1, 2] or []).
    """
    wb = _open_workbook_readonly(file_path)
    if wb is None:
        return []

    try:
        # Find the right sheet
        prefix = group_code[:2]
        course = COURSE_PREFIX_MAP.get(prefix)
        if course is None:
            return []

        sheet_name = resolve_sheet_name(wb, course)
        if sheet_name is None or sheet_name not in wb.sheetnames:
            return []

        ws = wb[sheet_name]

        # Find group columns
        columns = find_group_columns(ws, group_code)
        if columns is None:
            return []

        subject_col, _, _ = columns
        from hse_parser.reader import _col_to_index

        subj_idx = _col_to_index(subject_col)
        subgroups: set[int] = set()

        for row_cells in ws.iter_rows(min_row=12, values_only=False):
            if subj_idx >= len(row_cells):
                continue
            cell = row_cells[subj_idx]
            if cell.value is not None:
                text = str(cell.value)
                # Look for "гр.1" or "гр.2" patterns
                for match in re.finditer(r"гр\.?\s*(\d+)", text):
                    sg = int(match.group(1))
                    if sg in (1, 2):
                        subgroups.add(sg)

        return sorted(subgroups)
    finally:
        wb.close()


def suggest_academic_year() -> int:
    """Suggest the academic year start based on current date.

    If current month is >= September, academic year started this year.
    Otherwise, it started last year.
    """
    from datetime import date

    today = date.today()
    if today.month >= 9:
        return today.year
    return today.year - 1


def _extract_academic_year_from_title(title: str) -> int | None:
    """Extract the academic year start from the calendar title.

    E.g. 'КАЛЕНДАРЬ НА 2025-2026 учебный год' → 2025
    """
    match = re.search(r"(\d{4})", title)
    if match:
        return int(match.group(1))
    return None


def _read_upper_week_fill(ws: openpyxl.Worksheet) -> str | None:
    """Read the legend rows (3-5) to find the fill colour for 'верхняя неделя'.

    Returns the RGB colour string (e.g. 'FFFFE1CC') or None if not found.
    This approach is robust across different Excel files where colours may vary.
    """
    for row_num in (3, 4, 5):
        cell = ws.cell(row=row_num, column=1)
        if cell.value is not None:
            text = str(cell.value).strip().lower()
            if "верхн" in text:
                fill = cell.fill
                fg_color = fill.fgColor
                if fg_color and fg_color.rgb and fg_color.rgb != "00000000":
                    return fg_color.rgb
    return None


def _build_month_column_map(
    ws: openpyxl.Worksheet, header_row: int, max_col: int
) -> dict[int, int]:
    """Build a mapping of column index → month number for a module block.

    Reads the header row to find month names and their column ranges.
    Returns dict: {column_index: month_number}
    """
    col_to_month: dict[int, int] = {}
    month_boundaries: list[tuple[int, int]] = []  # (col_index, month_number)

    for col in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col)
        if cell.value is not None:
            name = str(cell.value).strip()
            if name in CALENDAR_MONTHS:
                month_boundaries.append((col, CALENDAR_MONTHS[name]))

    if not month_boundaries:
        return col_to_month

    # Fill in column ranges for each month
    for i, (start_col, month_num) in enumerate(month_boundaries):
        if i + 1 < len(month_boundaries):
            end_col = month_boundaries[i + 1][0] - 1
        else:
            end_col = max_col
        for c in range(start_col, end_col + 1):
            col_to_month[c] = month_num

    return col_to_month


def parse_calendar_sheet(file_path: str | Path) -> frozenset[date] | None:
    """Parse the 'Календарь' sheet to determine week parity by cell fill colour.

    Reads the calendar grid legend to dynamically determine which fill colour
    corresponds to 'верхняя неделя', then scans all module blocks to collect
    upper-week dates.

    Returns a frozenset of upper-week date objects, or None if the sheet
    doesn't exist or can't be read.
    """
    file_path = Path(file_path)
    if not file_path.exists():
        return None

    try:
        # Must use non-read_only to access cell fill colours and max_column
        wb = openpyxl.load_workbook(str(file_path))
    except Exception:
        return None

    try:
        if "Календарь" not in wb.sheetnames:
            return None

        ws = wb["Календарь"]

        # Extract academic year from title
        title_cell = ws.cell(row=1, column=1).value
        if title_cell is None:
            return None

        ac_year = _extract_academic_year_from_title(str(title_cell))
        if ac_year is None:
            return None

        # Read the upper-week fill colour from the legend (rows 3-5)
        upper_fill = _read_upper_week_fill(ws)
        if upper_fill is None:
            return None

        max_col = ws.max_column or 0
        if max_col < FIRST_DATA_COL:
            return None

        upper_dates: set[date] = set()

        for block_start in MODULE_BLOCK_ROWS:
            header_row = block_start + 1

            # Build month column mapping for this module block
            col_to_month = _build_month_column_map(ws, header_row, max_col)
            if not col_to_month:
                continue

            # Process each day-of-week row
            for offset in DAY_ROW_OFFSETS:
                row_num = block_start + offset
                if row_num > (ws.max_row or 0):
                    continue

                for col in range(FIRST_DATA_COL, max_col + 1):
                    cell = ws.cell(row=row_num, column=col)
                    if cell.value is None:
                        continue

                    month_num = col_to_month.get(col)
                    if month_num is None:
                        continue

                    # Determine year: Sept-Dec use ac_year, Jan-Aug use ac_year + 1
                    if month_num >= 9:
                        year = ac_year
                    else:
                        year = ac_year + 1

                    # Check if value is a number (date)
                    try:
                        day_num = int(float(str(cell.value)))
                    except (ValueError, TypeError):
                        continue

                    if not (1 <= day_num <= 31):
                        continue

                    # Check fill colour against the dynamically-read legend colour
                    fill = cell.fill
                    fg_color = fill.fgColor
                    if fg_color and fg_color.rgb and fg_color.rgb != "00000000":
                        if fg_color.rgb == upper_fill:
                            try:
                                upper_dates.add(date(year, month_num, day_num))
                            except ValueError:
                                continue

        return frozenset(upper_dates)
    finally:
        wb.close()