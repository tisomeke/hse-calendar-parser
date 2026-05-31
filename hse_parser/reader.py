"""Excel file reader for HSE schedule parser.

Reads .xlsx files, selects the correct sheet based on group code,
extracts module period, finds group columns, and iterates schedule rows.
Also reads the 'Календарь' sheet for week parity information.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import date, time

import openpyxl

from hse_parser.utils import (
    extract_module_from_title,
    extract_period_from_title,
    normalize_group_code,
    parse_time_range,
    resolve_year,
)

logger = logging.getLogger(__name__)

# Map group code prefix (first 2 digits) to course number
COURSE_PREFIX_MAP: dict[str, int] = {
    "25": 1,
    "24": 2,
    "23": 3,
    "22": 4,
}

# Map course number to sheet name
COURSE_TO_SHEET: dict[int, str] = {
    1: "1 курс",
    2: "2 курс",
    3: "3 курс",
    4: "4 курс",
}


@dataclass
class ScheduleRow:
    """A single row from the schedule table."""

    day_name: str  # Russian day name
    time_range: str  # e.g. "08:00-09:20"
    subject_text: str  # Raw cell text for the group
    auditorium: str  # Raw auditorium cell text
    building: str  # Raw building cell text
    row_number: int  # Excel row number
    col_letter: str  # Column letter of the subject cell


@dataclass
class ReaderResult:
    """Result from the reader module."""

    module: int  # Module number (3 or 4)
    module_period_start: str  # e.g. "09.01"
    module_period_end: str  # e.g. "24.03"
    academic_year_start: int
    rows: list[ScheduleRow] = field(default_factory=list)
    calendar_upper_dates: frozenset[date] | None = None
    """Upper-week dates from the 'Календарь' sheet, or None if unavailable."""


def find_sheet_name(wb: openpyxl.Workbook, group_code: str) -> str | None:
    """Find the correct sheet name for a given group code.

    Uses the group code prefix to determine the course,
    then finds the matching sheet. For 1st year, prefers
    '1 курс.' (current module) over '1 курс' (module 2).
    Falls back to the alternative sheet if the group is not
    found in the preferred one.
    """
    prefix = group_code[:2]
    course = COURSE_PREFIX_MAP.get(prefix)
    if course is None:
        return None

    base_name = COURSE_TO_SHEET[course]
    available = wb.sheetnames

    # For 1st year, prefer "1 курс." (current module)
    if course == 1:
        preferred = "1 курс."
        alternative = "1 курс"
        if preferred in available:
            # Check if group exists in preferred sheet
            ws = wb[preferred]
            if _group_exists_in_sheet(ws, group_code):
                return preferred
            # Fall back to alternative
            if alternative in available:
                return alternative
            return None
        if alternative in available:
            return alternative
        return None

    if base_name in available:
        return base_name
    return None


def _group_exists_in_sheet(ws: openpyxl.Worksheet, group_code: str) -> bool:
    """Check if a group code exists in the sheet's row 10."""
    from hse_parser.utils import normalize_group_code
    normalized = normalize_group_code(group_code)
    for row in ws.iter_rows(min_row=10, max_row=10, values_only=False):
        for cell in row:
            if cell.value is not None:
                cell_val = str(cell.value).strip()
                if normalize_group_code(cell_val) == normalized:
                    return True
    return False


def find_group_columns(
    ws: openpyxl.Worksheet, group_code: str
) -> tuple[str, str, str] | None:
    """Find the column letters for a group's subject, auditorium, and building.

    Searches row 10 (0-indexed: row 10 in Excel) for the group code.
    Returns (subject_col, auditorium_col, building_col) or None.

    The pattern is: [Group][Ауд.][Корпус] in consecutive columns.
    """
    normalized = normalize_group_code(group_code)

    for row in ws.iter_rows(min_row=10, max_row=10, values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            cell_val = str(cell.value).strip()
            if normalize_group_code(cell_val) == normalized:
                col = cell.column_letter
                # Next column should be Ауд., then Корпус
                next_col = _next_col(col)
                next_next_col = _next_col(next_col)
                return (col, next_col, next_next_col)
    return None


def _next_col(col: str) -> str:
    """Get the next Excel column letter.

    A → B, Z → AA, etc.
    """
    if not col:
        return "A"
    # Simple single-letter case
    if len(col) == 1:
        if col == "Z":
            return "AA"
        return chr(ord(col) + 1)
    # Multi-letter: increment like Excel
    chars = list(col)
    for i in range(len(chars) - 1, -1, -1):
        if chars[i] != "Z":
            chars[i] = chr(ord(chars[i]) + 1)
            break
        chars[i] = "A"
    else:
        chars.insert(0, "A")
    return "".join(chars)


def _col_to_index(col: str) -> int:
    """Convert Excel column letter to 0-based index."""
    result = 0
    for c in col:
        result = result * 26 + (ord(c) - ord("A") + 1)
    return result - 1


def extract_title_info(
    ws: openpyxl.Worksheet,
) -> tuple[int, str, str] | None:
    """Extract module number and period from the title row (row 1).

    Returns (module_number, period_start_str, period_end_str) or None.
    """
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            title = str(cell.value).strip()
            module = extract_module_from_title(title)
            period = extract_period_from_title(title)
            if module is not None and period is not None:
                return (module, period[0], period[1])
    return None


def read_schedule(
    file_path: str, group_code: str, academic_year_start: int = 2025
) -> ReaderResult | None:
    """Read schedule data from an Excel file for a specific group.

    Also reads the 'Календарь' sheet for week parity information.

    Returns a ReaderResult with all parsed rows, or None if the
    group/sheet cannot be found.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True)

    try:
        sheet_name = find_sheet_name(wb, group_code)
        if sheet_name is None:
            logger.error(
                "Could not find sheet for group %s. Available sheets: %s",
                group_code,
                wb.sheetnames,
            )
            return None

        ws = wb[sheet_name]

        # Extract module info from title
        title_info = extract_title_info(ws)
        if title_info is None:
            logger.error("Could not extract module info from sheet %s", sheet_name)
            return None

        module, period_start, period_end = title_info

        # Find group columns
        columns = find_group_columns(ws, group_code)
        if columns is None:
            logger.error(
                "Could not find group %s in sheet %s", group_code, sheet_name
            )
            return None

        subject_col, aud_col, build_col = columns

        # Iterate schedule rows (starting from row 12)
        rows: list[ScheduleRow] = []
        current_day: str | None = None

        for row_cells in ws.iter_rows(
            min_row=12, values_only=False
        ):
            row_num = row_cells[0].row

            # Check for footer markers
            has_footer = False
            for cell in row_cells:
                if cell.value is not None:
                    val = str(cell.value).strip()
                    if "Физическая культура" in val:
                        has_footer = True
                        break
            if has_footer:
                break

            # Get day name (column B, index 1)
            day_cell = row_cells[1] if len(row_cells) > 1 else None
            if day_cell and day_cell.value is not None:
                day_val = str(day_cell.value).strip()
                if day_val and day_val.lower() in (
                    "понедельник",
                    "вторник",
                    "среда",
                    "четверг",
                    "пятница",
                    "суббота",
                    "воскресенье",
                ):
                    current_day = day_val

            # Get time (column C, index 2)
            time_cell = row_cells[2] if len(row_cells) > 2 else None
            time_val = ""
            if time_cell and time_cell.value is not None:
                time_val = str(time_cell.value).strip()

            # Skip rows without time
            if not time_val:
                continue

            # Get subject cell
            subj_idx = _col_to_index(subject_col)
            aud_idx = _col_to_index(aud_col)
            build_idx = _col_to_index(build_col)

            subj_cell = row_cells[subj_idx] if subj_idx < len(row_cells) else None
            aud_cell = row_cells[aud_idx] if aud_idx < len(row_cells) else None
            build_cell = (
                row_cells[build_idx] if build_idx < len(row_cells) else None
            )

            subj_text = ""
            if subj_cell and subj_cell.value is not None:
                subj_text = str(subj_cell.value)

            aud_text = ""
            if aud_cell and aud_cell.value is not None:
                aud_text = str(aud_cell.value)

            build_text = ""
            if build_cell and build_cell.value is not None:
                build_text = str(build_cell.value)

            if current_day is None:
                continue

            rows.append(
                ScheduleRow(
                    day_name=current_day,
                    time_range=time_val,
                    subject_text=subj_text,
                    auditorium=aud_text,
                    building=build_text,
                    row_number=row_num,
                    col_letter=subject_col,
                )
            )

        # Read calendar sheet for week parity (lazy import to avoid circular deps)
        calendar_upper_dates: frozenset[date] | None = None
        try:
            from hse_schedule_parser.autodetect import parse_calendar_sheet
            calendar_upper_dates = parse_calendar_sheet(file_path)
        except Exception:
            logger.debug("Could not read calendar sheet for week parity")

        return ReaderResult(
            module=module,
            module_period_start=period_start,
            module_period_end=period_end,
            academic_year_start=academic_year_start,
            rows=rows,
            calendar_upper_dates=calendar_upper_dates,
        )

    finally:
        wb.close()