"""Tests for the autodetect module.

Tests cover:
- detect_courses: finding available course sheets
- detect_groups: finding group codes in a course sheet
- detect_subgroups: detecting subgroup markers
- detect_module_info: detecting module number and period
- suggest_academic_year: year suggestion logic
- parse_calendar_sheet: reading week parity from the 'Календарь' sheet
"""

from __future__ import annotations

import tempfile
from datetime import date
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from hse_schedule_parser.autodetect import (
    detect_courses,
    detect_groups,
    detect_module_info,
    detect_subgroups,
    parse_calendar_sheet,
    suggest_academic_year,
)


# ── Fixtures ────────────────────────────────────────────────────────────


@pytest.fixture
def workbook_with_courses() -> openpyxl.Workbook:
    """Create a workbook with standard course sheets."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name in ["1 курс", "2 курс", "3 курс", "4 курс"]:
        ws = wb.create_sheet(title=sheet_name)
        # Add group codes in row 10
        for col_idx, group in enumerate(["25ФПЛ1", "25ФПЛ2", "25БИВ1"], start=1):
            ws.cell(row=10, column=col_idx, value=group)

    return wb


@pytest.fixture
def workbook_with_subgroups() -> openpyxl.Workbook:
    """Create a workbook with subgroup markers in cells."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="1 курс")

    # Add group code in row 10
    ws.cell(row=10, column=1, value="25ФПЛ1")

    # Add subject cells with subgroup markers (row 12+)
    ws.cell(row=12, column=1, value="Иностранный язык (гр.1)")
    ws.cell(row=13, column=1, value="Иностранный язык (гр.2)")
    ws.cell(row=14, column=1, value="История (без подгруппы)")

    return wb


@pytest.fixture
def workbook_no_courses() -> openpyxl.Workbook:
    """Create a workbook with no recognizable course sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet(title="Sheet1")
    wb.create_sheet(title="Data")
    return wb


@pytest.fixture
def workbook_with_module_info() -> openpyxl.Workbook:
    """Create a workbook with module info in cell A1."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="1 курс")
    ws.cell(row=1, column=1, value="1 модуль (02.09 – 27.10)")
    ws.cell(row=10, column=1, value="25ФПЛ1")
    return wb


@pytest.fixture
def workbook_without_module_info() -> openpyxl.Workbook:
    """Create a workbook without module info in cell A1."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="1 курс")
    ws.cell(row=1, column=1, value="Расписание 1 курс")
    ws.cell(row=10, column=1, value="25ФПЛ1")
    return wb


@pytest.fixture
def temp_xlsx(workbook_with_courses) -> Path:
    """Save a workbook to a temp file and return the path."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)
    workbook_with_courses.save(str(path))
    yield path
    path.unlink(missing_ok=True)


@pytest.fixture
def temp_xlsx_subgroups(workbook_with_subgroups) -> Path:
    """Save a workbook with subgroups to a temp file."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)
    workbook_with_subgroups.save(str(path))
    yield path
    path.unlink(missing_ok=True)


@pytest.fixture
def temp_xlsx_no_courses(workbook_no_courses) -> Path:
    """Save a workbook with no courses to a temp file."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)
    workbook_no_courses.save(str(path))
    yield path
    path.unlink(missing_ok=True)


@pytest.fixture
def temp_xlsx_module(workbook_with_module_info) -> Path:
    """Save a workbook with module info to a temp file."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)
    workbook_with_module_info.save(str(path))
    yield path
    path.unlink(missing_ok=True)


@pytest.fixture
def temp_xlsx_no_module(workbook_without_module_info) -> Path:
    """Save a workbook without module info to a temp file."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)
    workbook_without_module_info.save(str(path))
    yield path
    path.unlink(missing_ok=True)


# ── detect_courses ──────────────────────────────────────────────────────


class TestDetectCourses:
    """Tests for detect_courses()."""

    def test_detects_all_courses(self, temp_xlsx: Path) -> None:
        """Should detect courses 1-4 when all sheets present."""
        courses = detect_courses(temp_xlsx)
        assert courses == [1, 2, 3, 4]

    def test_returns_empty_for_missing_file(self) -> None:
        """Should return empty list for non-existent file."""
        courses = detect_courses("/nonexistent/file.xlsx")
        assert courses == []

    def test_returns_empty_for_no_courses(self, temp_xlsx_no_courses: Path) -> None:
        """Should return empty list when no course sheets found."""
        courses = detect_courses(temp_xlsx_no_courses)
        assert courses == []

    def test_handles_corrupted_file(self) -> None:
        """Should return empty list for corrupted file."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(b"not an excel file")
            path = Path(f.name)
        try:
            courses = detect_courses(path)
            assert courses == []
        finally:
            path.unlink(missing_ok=True)

    def test_detects_single_course(self) -> None:
        """Should detect just one course if only one sheet."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet(title="1 курс")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)
        wb.save(str(path))
        try:
            courses = detect_courses(path)
            assert courses == [1]
        finally:
            path.unlink(missing_ok=True)

    def test_detects_first_course_variant(self) -> None:
        """Should detect '1 курс.' variant for first year."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet(title="1 курс.")
        wb.create_sheet(title="2 курс")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)
        wb.save(str(path))
        try:
            courses = detect_courses(path)
            assert 1 in courses
            assert 2 in courses
        finally:
            path.unlink(missing_ok=True)

    def test_returns_sorted_courses(self) -> None:
        """Should return courses in sorted order."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet(title="3 курс")
        wb.create_sheet(title="1 курс")
        wb.create_sheet(title="2 курс")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)
        wb.save(str(path))
        try:
            courses = detect_courses(path)
            assert courses == [1, 2, 3]
        finally:
            path.unlink(missing_ok=True)

    def test_ignores_non_course_sheets(self) -> None:
        """Should ignore sheets that don't match course pattern."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet(title="1 курс")
        wb.create_sheet(title="Лист1")
        wb.create_sheet(title="Данные")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)
        wb.save(str(path))
        try:
            courses = detect_courses(path)
            assert courses == [1]
        finally:
            path.unlink(missing_ok=True)


# ── detect_groups ───────────────────────────────────────────────────────


class TestDetectGroups:
    """Tests for detect_groups()."""

    def test_detects_groups(self, temp_xlsx: Path) -> None:
        """Should detect group codes for a course."""
        groups = detect_groups(temp_xlsx, 1)
        assert "25ФПЛ1" in groups
        assert "25ФПЛ2" in groups
        assert "25БИВ1" in groups

    def test_returns_empty_for_missing_file(self) -> None:
        """Should return empty list for non-existent file."""
        groups = detect_groups("/nonexistent/file.xlsx", 1)
        assert groups == []

    def test_returns_empty_for_nonexistent_course(self, temp_xlsx: Path) -> None:
        """Should return empty list for course not in file."""
        groups = detect_groups(temp_xlsx, 5)
        assert groups == []

    def test_returns_sorted_unique(self, temp_xlsx: Path) -> None:
        """Should return sorted unique group codes."""
        groups = detect_groups(temp_xlsx, 1)
        assert groups == sorted(set(groups))

    def test_handles_corrupted_file(self) -> None:
        """Should return empty list for corrupted file."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(b"corrupted data")
            path = Path(f.name)
        try:
            groups = detect_groups(path, 1)
            assert groups == []
        finally:
            path.unlink(missing_ok=True)

    def test_returns_empty_for_empty_sheet(self) -> None:
        """Should return empty list when sheet has no data."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet(title="1 курс")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)
        wb.save(str(path))
        try:
            groups = detect_groups(path, 1)
            assert groups == []
        finally:
            path.unlink(missing_ok=True)

    def test_detects_groups_in_row_10_only(self) -> None:
        """Should only look at row 10 for group codes."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title="1 курс")
        ws.cell(row=9, column=1, value="NOT_GROUP")
        ws.cell(row=10, column=1, value="25ФПЛ1")
        ws.cell(row=11, column=1, value="NOT_GROUP2")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)
        wb.save(str(path))
        try:
            groups = detect_groups(path, 1)
            assert groups == ["25ФПЛ1"]
        finally:
            path.unlink(missing_ok=True)

    def test_filters_non_group_codes(self) -> None:
        """Should filter out cells that don't look like group codes."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title="1 курс")
        ws.cell(row=10, column=1, value="25ФПЛ1")
        ws.cell(row=10, column=2, value="NOT_A_GROUP")
        ws.cell(row=10, column=3, value="")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)
        wb.save(str(path))
        try:
            groups = detect_groups(path, 1)
            assert groups == ["25ФПЛ1"]
        finally:
            path.unlink(missing_ok=True)


# ── detect_subgroups ────────────────────────────────────────────────────


class TestDetectSubgroups:
    """Tests for detect_subgroups()."""

    def test_detects_subgroups(self, temp_xlsx_subgroups: Path) -> None:
        """Should detect subgroups 1 and 2 from cell content."""
        subgroups = detect_subgroups(temp_xlsx_subgroups, "25ФПЛ1")
        assert 1 in subgroups
        assert 2 in subgroups

    def test_returns_empty_for_no_subgroups(self, temp_xlsx: Path) -> None:
        """Should return empty list when no subgroup markers."""
        subgroups = detect_subgroups(temp_xlsx, "25ФПЛ1")
        assert subgroups == []

    def test_returns_empty_for_missing_file(self) -> None:
        """Should return empty list for non-existent file."""
        subgroups = detect_subgroups("/nonexistent/file.xlsx", "25ФПЛ1")
        assert subgroups == []

    def test_returns_empty_for_unknown_prefix(self, temp_xlsx: Path) -> None:
        """Should return empty list for unknown group prefix."""
        subgroups = detect_subgroups(temp_xlsx, "99XXXX")
        assert subgroups == []

    def test_detects_subgroups_with_group_variant(self) -> None:
        """Should detect subgroups with 'гр. 1' format."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title="1 курс")
        ws.cell(row=10, column=1, value="25ФПЛ1")
        ws.cell(row=12, column=1, value="Физика (гр. 1)")
        ws.cell(row=13, column=1, value="Физика (гр. 2)")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)
        wb.save(str(path))
        try:
            subgroups = detect_subgroups(path, "25ФПЛ1")
            assert 1 in subgroups
            assert 2 in subgroups
        finally:
            path.unlink(missing_ok=True)

    def test_handles_corrupted_file(self) -> None:
        """Should return empty list for corrupted file."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(b"corrupted")
            path = Path(f.name)
        try:
            subgroups = detect_subgroups(path, "25ФПЛ1")
            assert subgroups == []
        finally:
            path.unlink(missing_ok=True)


# ── detect_module_info ──────────────────────────────────────────────────


class TestDetectModuleInfo:
    """Tests for detect_module_info()."""

    def test_returns_none_without_module_info(self, temp_xlsx_no_module: Path) -> None:
        """Should return None when no module info in cell A1."""
        info = detect_module_info(temp_xlsx_no_module, 1)
        assert info is None

    def test_returns_none_for_missing_file(self) -> None:
        """Should return None for non-existent file."""
        info = detect_module_info("/nonexistent/file.xlsx", 1)
        assert info is None

    def test_returns_none_for_nonexistent_course(self, temp_xlsx_module: Path) -> None:
        """Should return None for course not in file."""
        info = detect_module_info(temp_xlsx_module, 5)
        assert info is None

    def test_detects_second_module(self) -> None:
        """Should detect module 2."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title="1 курс")
        ws.cell(row=1, column=1, value="2 модуль (11.11 – 29.12)")
        ws.cell(row=10, column=1, value="25ФПЛ1")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)
        wb.save(str(path))
        try:
            info = detect_module_info(path, 1)
            assert info is not None
            assert info["module"] == 2
        finally:
            path.unlink(missing_ok=True)

    def test_detects_module_with_different_date_format(self) -> None:
        """Should detect module with different date separator."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title="1 курс")
        ws.cell(row=1, column=1, value="3 модуль (10.01-23.03)")
        ws.cell(row=10, column=1, value="25ФПЛ1")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)
        wb.save(str(path))
        try:
            info = detect_module_info(path, 1)
            assert info is not None
            assert info["module"] == 3
        finally:
            path.unlink(missing_ok=True)

    def test_handles_corrupted_file(self) -> None:
        """Should return None for corrupted file."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(b"corrupted")
            path = Path(f.name)
        try:
            info = detect_module_info(path, 1)
            assert info is None
        finally:
            path.unlink(missing_ok=True)

    def test_returns_none_for_empty_cell_a1(self) -> None:
        """Should return None when cell A1 is empty."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title="1 курс")
        ws.cell(row=1, column=1, value=None)
        ws.cell(row=10, column=1, value="25ФПЛ1")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)
        wb.save(str(path))
        try:
            info = detect_module_info(path, 1)
            assert info is None
        finally:
            path.unlink(missing_ok=True)


# ── suggest_academic_year ───────────────────────────────────────────────


class TestSuggestAcademicYear:
    """Tests for suggest_academic_year()."""

    @patch("datetime.date")
    def test_suggests_current_year_in_september(self, mock_date: MagicMock) -> None:
        """Should suggest current year when month >= 9."""
        mock_date.today.return_value = type("Date", (), {"month": 9, "year": 2025})()
        assert suggest_academic_year() == 2025

    @patch("datetime.date")
    def test_suggests_previous_year_in_march(self, mock_date: MagicMock) -> None:
        """Should suggest previous year when month < 9."""
        mock_date.today.return_value = type("Date", (), {"month": 3, "year": 2025})()
        assert suggest_academic_year() == 2024

    @patch("datetime.date")
    def test_suggests_previous_year_in_january(self, mock_date: MagicMock) -> None:
        """Should suggest previous year in January."""
        mock_date.today.return_value = type("Date", (), {"month": 1, "year": 2025})()
        assert suggest_academic_year() == 2024

    @patch("datetime.date")
    def test_suggests_current_year_in_december(self, mock_date: MagicMock) -> None:
        """Should suggest current year in December."""
        mock_date.today.return_value = type("Date", (), {"month": 12, "year": 2025})()
        assert suggest_academic_year() == 2025

    @patch("datetime.date")
    def test_suggests_current_year_in_august(self, mock_date: MagicMock) -> None:
        """Should suggest previous year in August (month < 9)."""
        mock_date.today.return_value = type("Date", (), {"month": 8, "year": 2025})()
        assert suggest_academic_year() == 2024

    @patch("datetime.date")
    def test_suggests_current_year_in_june(self, mock_date: MagicMock) -> None:
        """Should suggest previous year in June."""
        mock_date.today.return_value = type("Date", (), {"month": 6, "year": 2025})()
        assert suggest_academic_year() == 2024

    @patch("datetime.date")
    def test_suggests_current_year_in_october(self, mock_date: MagicMock) -> None:
        """Should suggest current year in October."""
        mock_date.today.return_value = type("Date", (), {"month": 10, "year": 2025})()
        assert suggest_academic_year() == 2025

    @patch("datetime.date")
    def test_edge_year_2024(self, mock_date: MagicMock) -> None:
        """Should work with year 2024."""
        mock_date.today.return_value = type("Date", (), {"month": 9, "year": 2024})()
        assert suggest_academic_year() == 2024

    @patch("datetime.date")
    def test_edge_year_2030(self, mock_date: MagicMock) -> None:
        """Should work with year 2030."""
        mock_date.today.return_value = type("Date", (), {"month": 3, "year": 2030})()
        assert suggest_academic_year() == 2029


# ── Fixtures for parse_calendar_sheet ─────────────────────────────────────


@pytest.fixture
def calendar_workbook() -> openpyxl.Workbook:
    """Create a minimal workbook with a 'Календарь' sheet.

    Structure mimics the real file:
    - Row 1: title with academic year
    - Row 3: legend 'верхняя неделя' with peach fill
    - Row 4: legend 'сессия' with blue fill
    - Row 5: legend 'каникулы' with green fill
    - Module I block (rows 7-16): September-October
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Календарь"

    # Title
    ws.cell(row=1, column=1, value="КАЛЕНДАРЬ НА 2025-2026 учебный год")

    # Legend
    from openpyxl.styles import PatternFill
    peach = PatternFill(start_color="FFFFE1CC", end_color="FFFFE1CC", fill_type="solid")
    blue = PatternFill(start_color="FF8CB5F9", end_color="FF8CB5F9", fill_type="solid")
    green = PatternFill(start_color="FF1A5529", end_color="FF1A5529", fill_type="solid")

    ws.cell(row=3, column=1, value="верхняя неделя").fill = peach
    ws.cell(row=4, column=1, value="сессия").fill = blue
    ws.cell(row=5, column=1, value="каникулы").fill = green

    # Module I header
    ws.cell(row=7, column=1, value="I модуль")
    ws.cell(row=8, column=2, value="Дни недели")
    ws.cell(row=8, column=4, value="Сентябрь")
    ws.cell(row=8, column=10, value="Октябрь")

    # Day rows for Module I (rows 10-16)
    # Sept 1 (Mon) = upper, Sept 8 (Mon) = lower, Sept 15 (Mon) = upper
    # Oct 6 (Tue) = upper, Oct 13 (Tue) = lower
    days = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота", "воскресенье"]
    for i, day_name in enumerate(days):
        row_num = 10 + i
        ws.cell(row=row_num, column=2, value=day_name)
        # September dates (cols D-H = 4-8)
        ws.cell(row=row_num, column=4, value=1 + i)  # Sept 1-7
        ws.cell(row=row_num, column=5, value=8 + i)  # Sept 8-14
        ws.cell(row=row_num, column=6, value=15 + i)  # Sept 15-21
        ws.cell(row=row_num, column=7, value=22 + i)  # Sept 22-28
        ws.cell(row=row_num, column=8, value=29 + i)  # Sept 29-35 (overflow)
        # October dates (cols J-N = 10-14)
        ws.cell(row=row_num, column=10, value=6 + i)  # Oct 6-12
        ws.cell(row=row_num, column=11, value=13 + i)  # Oct 13-19
        ws.cell(row=row_num, column=12, value=20 + i)  # Oct 20-26
        ws.cell(row=row_num, column=13, value=27 + i)  # Oct 27-33 (overflow)

    # Apply peach fill to upper week cells (Sept 1-7, Sept 15-21, Oct 6-12)
    for i in range(7):
        # Sept 1-7 (col D)
        ws.cell(row=10 + i, column=4).fill = peach
        # Sept 15-21 (col F)
        ws.cell(row=10 + i, column=6).fill = peach
        # Oct 6-12 (col J)
        ws.cell(row=10 + i, column=10).fill = peach

    return wb


@pytest.fixture
def calendar_xlsx(calendar_workbook: openpyxl.Workbook) -> Path:
    """Save the calendar workbook to a temp file."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)
    calendar_workbook.save(str(path))
    calendar_workbook.close()
    yield path
    path.unlink(missing_ok=True)


# ── parse_calendar_sheet ──────────────────────────────────────────────────


class TestParseCalendarSheet:
    """Tests for parse_calendar_sheet()."""

    def test_returns_upper_dates(self, calendar_xlsx: Path) -> None:
        """Should return upper-week dates from the calendar sheet."""
        result = parse_calendar_sheet(calendar_xlsx)
        assert result is not None
        assert isinstance(result, frozenset)

        # Sept 1, 2025 (Monday) — upper (peach)
        assert date(2025, 9, 1) in result
        # Sept 2, 2025 (Tuesday) — upper (peach)
        assert date(2025, 9, 2) in result
        # Sept 8, 2025 (Monday) — lower (no fill)
        assert date(2025, 9, 8) not in result
        # Sept 15, 2025 (Monday) — upper (peach)
        assert date(2025, 9, 15) in result
        # Oct 6, 2025 (Monday) — upper (peach)
        assert date(2025, 10, 6) in result
        # Oct 13, 2025 (Monday) — lower (no fill)
        assert date(2025, 10, 13) not in result

    def test_returns_none_for_missing_file(self) -> None:
        """Should return None for non-existent file."""
        result = parse_calendar_sheet("/nonexistent/file.xlsx")
        assert result is None

    def test_returns_none_for_no_calendar_sheet(self, tmp_path: Path) -> None:
        """Should return None when workbook has no 'Календарь' sheet."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "1 курс"
        path = tmp_path / "no_calendar.xlsx"
        wb.save(str(path))
        wb.close()

        result = parse_calendar_sheet(path)
        assert result is None

    def test_returns_none_for_corrupted_file(self) -> None:
        """Should return None for corrupted file."""
        result = parse_calendar_sheet("/dev/null")
        assert result is None

    def test_returns_none_for_empty_title(self, tmp_path: Path) -> None:
        """Should return None when title cell is empty."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Календарь"
        # No title in A1
        path = tmp_path / "empty_title.xlsx"
        wb.save(str(path))
        wb.close()

        result = parse_calendar_sheet(path)
        assert result is None

    def test_returns_none_without_legend(self, tmp_path: Path) -> None:
        """Should return None when legend row is missing."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Календарь"
        ws.cell(row=1, column=1, value="КАЛЕНДАРЬ НА 2025-2026 учебный год")
        # No legend rows
        path = tmp_path / "no_legend.xlsx"
        wb.save(str(path))
        wb.close()

        result = parse_calendar_sheet(path)
        assert result is None

    def test_handles_different_academic_year(self, tmp_path: Path) -> None:
        """Should handle different academic year in title."""
        from openpyxl.styles import PatternFill
        peach = PatternFill(start_color="FFFFE1CC", end_color="FFFFE1CC", fill_type="solid")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Календарь"
        ws.cell(row=1, column=1, value="КАЛЕНДАРЬ НА 2024-2025 учебный год")
        ws.cell(row=3, column=1, value="верхняя неделя").fill = peach
        ws.cell(row=7, column=1, value="I модуль")
        ws.cell(row=8, column=4, value="Сентябрь")
        ws.cell(row=10, column=2, value="понедельник")
        ws.cell(row=10, column=4, value=1).fill = peach  # Sept 1, 2024

        path = tmp_path / "year_2024.xlsx"
        wb.save(str(path))
        wb.close()

        result = parse_calendar_sheet(path)
        assert result is not None
        assert date(2024, 9, 1) in result

    def test_handles_january_dates(self, tmp_path: Path) -> None:
        """Should use ac_year + 1 for January dates."""
        from openpyxl.styles import PatternFill
        peach = PatternFill(start_color="FFFFE1CC", end_color="FFFFE1CC", fill_type="solid")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Календарь"
        ws.cell(row=1, column=1, value="КАЛЕНДАРЬ НА 2025-2026 учебный год")
        ws.cell(row=3, column=1, value="верхняя неделя").fill = peach
        ws.cell(row=29, column=1, value="III модуль")
        ws.cell(row=30, column=4, value="Январь")
        ws.cell(row=32, column=2, value="понедельник")
        ws.cell(row=32, column=5, value=12).fill = peach  # Jan 12, 2026

        path = tmp_path / "january.xlsx"
        wb.save(str(path))
        wb.close()

        result = parse_calendar_sheet(path)
        assert result is not None
        # January should use ac_year + 1 = 2026
        assert date(2026, 1, 12) in result
        assert date(2025, 1, 12) not in result